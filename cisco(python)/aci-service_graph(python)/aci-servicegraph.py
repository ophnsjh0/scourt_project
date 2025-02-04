import requests
import os
import sys
import argparse
import json
import re
import pandas as pd
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook
import warnings
from urllib3.exceptions import InsecureRequestWarning

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
warnings.filterwarnings("ignore", category=UserWarning)

class ServiceGraph():

    def __init__(self):
        self.base_url   = os.environ['ACI_URL']
        self.username   = os.environ['ACI_USERNAME']
        self.password   = os.environ['ACI_PASSWORD']
        self.ssl_verify = False if 'ACI_SSLVERIFY' in os.environ and os.environ['ACI_SSLVERIFY'] == "False" else True
        self.headers    = {"Content-Type": "application/json"}
        self.filepath   = 'request_service_graph.xlsx'
        
        env = Environment(loader=FileSystemLoader("templates/"))

        self.templetes                           = {}
        self.templetes['Filters']                = env.get_template("01_Filters.j2")
        self.templetes['Contracts']              = env.get_template("02_contracts.j2")
        self.templetes['Logical_Devices']        = env.get_template("03_logical_devices.j2")
        self.templetes['Service_Graph_Template'] = env.get_template("04_service_graph_template.j2")
        self.templetes['Apply_Service_Graph']    = env.get_template("05_apply_service_graph.j2")
        self.templetes['IP_SLA']                 = env.get_template("06_ip_sla.j2")
        
    def auth(self):
        payload = f'{{"aaaUser": {{ "attributes": {{ "name": "{ self.username }", "pwd": "{ self.password }" }} }} }}'
        response = requests.post(f'{self.base_url}/api/aaaLogin.json', headers=self.headers, data=payload, verify=self.ssl_verify)
        if response.status_code == 200:
            self.headers['Cookie'] = f"APIC-cookie={response.json()['imdata'][0]['aaaLogin']['attributes']['token']}"
        else:
            raise Exception("ACI auth failed.")
        return response.json()['imdata'][0]['aaaLogin']['attributes']

    def get(self, api):
        url = f'{self.base_url}/{api}'
        res = requests.get(url=url, headers=self.headers, verify=self.ssl_verify)
        return res

    def post(self, body):
        url = f'{self.base_url}/api/node/mo.json'
        res = requests.post(url=url, headers=self.headers, verify=self.ssl_verify, data=json.dumps(body))
        return res.status_code, json.dumps(body)

    def executeTask(self, taskName, processName=''):
        print(f'\n{processName if processName != None and processName != "" else taskName}')
        result = []
        isStop = False
        workbook = load_workbook(filename=self.filepath)
        sheet = workbook[taskName]
        header = [cell.value for cell in sheet[2]]
        cur_data_num = 1
        for row in sheet.iter_rows(min_row=3, values_only=True):
            row_dict = dict(zip(header, row))
            content = self.templetes[taskName].render(row_dict)
            # print(content)
            res = self.post(body=json.loads(content))
            if res[0] == 200 or res[0] == 201:
                print(f'  - attempting data #{cur_data_num} --> success')
            else:
                print(f'  - attempting data #{cur_data_num} --> fail')
                isStop = True
            cur_data_num = cur_data_num + 1
        if isStop:
            print('FAILED AND PROGRAM STOPPED! PLEASE CHECK DATA!!')
            exit(1)
    
    def checkGraphInstances(self):
        print(f'\nCheck Service Graph Instances deployed:')
        workbook = load_workbook(filename=self.filepath)
        sheet = workbook['Apply_Service_Graph']
        header = [cell.value for cell in sheet[2]]
        cur_data_num = 1
        for row in sheet.iter_rows(min_row=3, values_only=True):
            row_dict = dict(zip(header, row))
            uri = f'api/node/mo/uni/tn-{row_dict["TENANT"]}/GraphInst_C-[uni/tn-{row_dict["TENANT"]}/brc-{row_dict["CONTRACT"]}]-G-[uni/tn-{row_dict["TENANT"]}/AbsGraph-{row_dict["SERVICE_GRAPH_TEMPLATE"]}]-S-[uni/tn-{row_dict["TENANT"]}].json'
            res = self.get(api=uri)
            if res.status_code == 200:
                print(f"  - Applying Service Graph #{cur_data_num} --> success")
            else:
                print(f"  - Applying Service Graph #{cur_data_num} --> fail")

if __name__ == "__main__":
    print('''
=============================================================================
ACI Service Graph Builder - ver 0.5.0
Copyright (c) 2024 Cisco [Customer Success] and/or its affiliates.
=============================================================================

Excutable jobs:
    1. Configure IP SLA Monitoring Policy, Track memeber and Track list [IP_SLA]
    2. Configure filters [Filters]
    3. Configure contracts [Contracts]
    4. Configure Logical devices and Service Graph Template [Logical_Devices, Service_Graph_Template]
    5. Apply Service Graph Instances [Apply_Service_Graph]
''')
    params = input('Enter job numbers to execute [ example: 2,3,4 ]: ')

    svc = ServiceGraph()
    svc.auth()

    for num in params.split(','):
        match num:
            case '1':
                svc.executeTask('IP_SLA', 'Configure IP SLA Monitoring Policy, Track memeber, Track list')

            case '2':
                svc.executeTask('Filters', 'Configure filters')
            
            case '3':
                svc.executeTask('Contracts', 'Configure contracts')

            case '4':
                svc.executeTask('Logical_Devices', 'Configure logical devices')
                svc.executeTask('Service_Graph_Template', 'Configure service graph templates')

            case '5':
                svc.executeTask('Apply_Service_Graph', 'Apply service graph instances')
                svc.checkGraphInstances()
